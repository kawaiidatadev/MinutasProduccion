# sub_menus/procesar_excel.py
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    ScatterChart,
    Reference,
    Series
)
from openpyxl.chart.label import DataLabelList
import os


def excel_pr(filepath):
    """
    Procesa el archivo Excel para mejorar su presentación y agregar gráficos
    """
    try:
        # Verificar que el archivo existe
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"No se encontró el archivo: {filepath}")

        # Leer el archivo Excel
        df = pd.read_excel(filepath)

        # Crear un nuevo archivo Excel procesado
        processed_filename = f"procesado_{os.path.basename(filepath)}"
        processed_path = os.path.join(os.path.dirname(filepath), processed_filename)

        # Crear un nuevo workbook para el archivo procesado
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Acuerdos"

        # Escribir los encabezados
        for c_idx, column in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=c_idx, value=column)

        # Escribir los datos
        for r_idx, row in enumerate(df.itertuples(), 2):  # Comenzar en fila 2 para los datos
            for c_idx, value in enumerate(row[1:], 1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        # ==============================================
        # 1. FORMATO DE LA TABLA PRINCIPAL
        # ==============================================

        # Ajustar anchos de columna
        column_widths = {
            'A': 15,  # id_acuerdo
            'B': 40,  # acuerdo
            'C': 25,  # responsables
            'D': 15,  # fecha_compromiso
            'E': 18,  # fecha_registro
            'F': 20,  # usuario_registra
            'G': 12,  # estatus
            'H': 15,  # fecha_estatus
            'I': 30,  # comentarios_cierre
            'J': 12,  # diferencia_dias
            'K': 40  # historial
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Estilo para el encabezado
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Estilo para las celdas de datos
        data_font = Font(name='Calibri', size=11)
        data_alignment = Alignment(vertical='top', wrap_text=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # Alternar colores de fila para mejor lectura
        for row in range(2, worksheet.max_row + 1):
            fill_color = 'FFFFFF' if row % 2 == 0 else 'E6E6E6'
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Congelar paneles (encabezados visibles al desplazar)
        worksheet.freeze_panes = 'A2'

        # ==============================================
        # 2. CREACIÓN DE GRÁFICOS EN HOJAS SEPARADAS
        # ==============================================

        # Gráfico 1: Acuerdos por estatus (Pie Chart)
        if 'estatus' in df.columns:
            status_counts = df['estatus'].value_counts()
            ws_status = workbook.create_sheet(title="Estatus Acuerdos")

            ws_status.append(['Estatus', 'Cantidad'])
            for status, count in status_counts.items():
                ws_status.append([status, count])

            pie = PieChart()
            labels = Reference(ws_status, min_col=1, min_row=2, max_row=len(status_counts) + 1)
            data = Reference(ws_status, min_col=2, min_row=1, max_row=len(status_counts) + 1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribución de Acuerdos por Estatus"

            # Posicionar el gráfico
            ws_status.add_chart(pie, "D2")

            # Formato de la hoja
            ws_status.column_dimensions['A'].width = 20
            ws_status.column_dimensions['B'].width = 15

        # Gráfico 2: Días de diferencia promedio por responsable (Bar Chart)
        if 'responsables' in df.columns and 'diferencia_dias' in df.columns:
            # Procesar responsables (tomar solo el primer nombre)
            df['primer_responsable'] = df['responsables'].str.split(',').str[0].str.strip()
            avg_days = df.groupby('primer_responsable')['diferencia_dias'].mean().sort_values()

            ws_responsables = workbook.create_sheet(title="Rendimiento Responsables")

            ws_responsables.append(['Responsable', 'Días Promedio'])
            for resp, days in avg_days.items():
                ws_responsables.append([resp, days])

            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.title = "Días Promedio para Cumplimiento"
            bar_chart.y_axis.title = "Días"
            bar_chart.x_axis.title = "Responsable"

            data = Reference(ws_responsables, min_col=2, min_row=1, max_row=len(avg_days) + 1)
            cats = Reference(ws_responsables, min_col=1, min_row=2, max_row=len(avg_days) + 1)

            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            bar_chart.shape = 4

            ws_responsables.add_chart(bar_chart, "D2")

            # Formato de la hoja
            ws_responsables.column_dimensions['A'].width = 25
            ws_responsables.column_dimensions['B'].width = 15

        # Gráfico 3: Evolución temporal de acuerdos (Line Chart)
        if 'fecha_registro' in df.columns:
            df['fecha_registro'] = pd.to_datetime(df['fecha_registro']).dt.date
            timeline = df['fecha_registro'].value_counts().sort_index()

            ws_timeline = workbook.create_sheet(title="Evolución Temporal")

            ws_timeline.append(['Fecha', 'Acuerdos Registrados'])
            for date, count in timeline.items():
                ws_timeline.append([date, count])

            line_chart = LineChart()
            line_chart.title = "Acuerdos Registrados por Fecha"
            line_chart.y_axis.title = "Cantidad"
            line_chart.x_axis.title = "Fecha"

            data = Reference(ws_timeline, min_col=2, min_row=1, max_row=len(timeline) + 1)
            cats = Reference(ws_timeline, min_col=1, min_row=2, max_row=len(timeline) + 1)

            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(cats)
            line_chart.style = 2

            ws_timeline.add_chart(line_chart, "D2")

            # Formato de la hoja
            ws_timeline.column_dimensions['A'].width = 15
            ws_timeline.column_dimensions['B'].width = 20

        # Gráfico 4: Relación entre días de diferencia y estatus (Scatter Chart)
        if 'diferencia_dias' in df.columns and 'estatus' in df.columns:
            ws_scatter = workbook.create_sheet(title="Días vs Estatus")

            # Preparar datos para el scatter plot
            scatter_data = df[['estatus', 'diferencia_dias']]
            scatter_data = scatter_data.sort_values('estatus')

            ws_scatter.append(['Estatus', 'Días de Diferencia'])
            for _, row in scatter_data.iterrows():
                ws_scatter.append([row['estatus'], row['diferencia_dias']])

            scatter_chart = ScatterChart()
            scatter_chart.title = "Relación Días de Diferencia vs Estatus"
            scatter_chart.y_axis.title = "Días de Diferencia"
            scatter_chart.x_axis.title = "Estatus"

            xvalues = Reference(ws_scatter, min_col=1, min_row=2, max_row=len(scatter_data) + 1)
            yvalues = Reference(ws_scatter, min_col=2, min_row=2, max_row=len(scatter_data) + 1)

            series = Series(yvalues, xvalues, title_from_data=False)
            scatter_chart.series.append(series)

            ws_scatter.add_chart(scatter_chart, "D2")

            # Formato de la hoja
            ws_scatter.column_dimensions['A'].width = 15
            ws_scatter.column_dimensions['B'].width = 20

        # Gráfico 5: Top 10 acuerdos con mayor diferencia de días (Bar Chart)
        if 'diferencia_dias' in df.columns and 'acuerdo' in df.columns:
            top_delayed = df.nlargest(10, 'diferencia_dias')[['acuerdo', 'diferencia_dias']]

            ws_top_delayed = workbook.create_sheet(title="Top 10 Atrasados")

            ws_top_delayed.append(['Acuerdo', 'Días de Diferencia'])
            for _, row in top_delayed.iterrows():
                ws_top_delayed.append([row['acuerdo'], row['diferencia_dias']])

            bar_chart = BarChart()
            bar_chart.type = "bar"
            bar_chart.style = 11
            bar_chart.title = "Top 10 Acuerdos con Mayor Diferencia"
            bar_chart.y_axis.title = "Días"

            data = Reference(ws_top_delayed, min_col=2, min_row=1, max_row=11)
            cats = Reference(ws_top_delayed, min_col=1, min_row=2, max_row=11)

            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)

            # Añadir etiquetas de datos
            bar_chart.dataLabels = DataLabelList()
            bar_chart.dataLabels.showVal = True

            ws_top_delayed.add_chart(bar_chart, "D2")

            # Formato de la hoja
            ws_top_delayed.column_dimensions['A'].width = 40
            ws_top_delayed.column_dimensions['B'].width = 20

        # ==============================================
        # 3. GUARDAR Y CERRAR EL ARCHIVO
        # ==============================================

        # Eliminar la hoja en blanco que se crea por defecto
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        # Guardar el archivo procesado
        workbook.save(processed_path)
        workbook.close()

        return processed_path

    except Exception as e:
        print(f"Error al procesar el Excel: {str(e)}")
        # Si hay error, devolver el archivo original
        return filepath