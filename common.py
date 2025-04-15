import tkinter as tk
import threading
from tkinter import ttk, messagebox
from datetime import datetime
import sqlite3
from difflib import SequenceMatcher
from tkcalendar import Calendar
import os
import time
import textwrap
from tkinter import font
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    ScatterChart,
    Reference,
    Series
)
from openpyxl.chart.label import DataLabelList
import os
import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import pandas as pd
from datetime import datetime
import os
import sys
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import sqlite3
import getpass
from tkcalendar import DateEntry  # Necesitarás instalar: pip install tkcalendar
import os
import platform
from ctypes import windll
import pygame
import pygame
import playsound
import pyttsx3
from PIL import Image, ImageTk
import fpdf
import threading
import time
import os
import sys
import locale
import win32gui
import win32con
import sqlite3
import time
import threading
import os
import sys
import winreg as reg
from plyer import notification
import pystray
from PIL import Image
import io
import base64
import win32event
import win32api
import win32con
import ctypes
import os
import sys
import io
import time
import sqlite3
import threading
import win32event
import win32api
import win32con
from win32com.shell import shell, shellcon
import win32service
import win32serviceutil
import winerror
from PIL import Image
import pystray
from plyer import notification
import ctypes






def verificar_imports():
    imports = [
        'tkinter',
        'sqlite3',
        'datetime',
        'threading',
        ('difflib', 'SequenceMatcher'),  # Ahora como tupla (módulo, submodulo)
        'tkcalendar',
        'os',
        'time',
        'textwrap',
        'pandas',
        'openpyxl',
        ('openpyxl.styles', 'Font'),
        ('openpyxl.styles', 'Alignment'),
        ('openpyxl.styles', 'Border'),
        ('openpyxl.styles', 'Side'),
        ('openpyxl.styles', 'PatternFill'),
        ('openpyxl.utils', 'get_column_letter'),
        ('openpyxl.chart', 'BarChart'),
        'sys',
        'subprocess',
        'getpass',
        'platform',
        ('ctypes', 'windll'),
        'pygame',
        'playsound',
        'pyttsx3',
        ('PIL', 'Image'),
        'fpdf'
    ]

    print("=== Verificación de imports ===")
    print("Este script verificará si todos los paquetes necesarios están instalados.")
    print("=" * 40)

    for imp in imports:
        try:
            if isinstance(imp, tuple):
                module, submodule = imp
                imported = __import__(module, fromlist=[submodule])
                getattr(imported, submodule)
                print(f"[✓] {module}.{submodule.ljust(25)} - Correctamente instalado")
            else:
                __import__(imp)
                print(f"[✓] {imp.ljust(30)} - Correctamente instalado")
        except ImportError as e:
            name = f"{imp[0]}.{imp[1]}" if isinstance(imp, tuple) else imp
            print(f"[✗] {name.ljust(30)} - FALTA INSTALAR: {str(e)}")
        except AttributeError as e:
            print(f"[✗] {imp[0]}.{imp[1].ljust(25)} - Submódulo no encontrado: {str(e)}")

    print("=" * 40)
    print("Nota: Los paquetes que muestran [✗] necesitan ser instalados.")
    print("Puedes instalarlos con: pip install nombre_paquete")




